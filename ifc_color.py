# After running this script, save it to a new file then open the new file you just saved to check the color result

from openpyxl import load_workbook
import ifcopenshell
import ifcopenshell.util.element
from bonsai.bim.ifc import IfcStore
import bonsai.tool as tool
import re
import bpy

# --------------------- CONFIG ---------------------
XLSX_PATH = r"/Users/jeanhu/Desktop/SBI/ifc_project/project2_color/SimpleBIM_Type_Filter.xlsx"
SHEET_NAME = "ModelView"
# --------------------------------------------------

def argb_to_rgb(argb_hex_string):
    # '#AARRGGBB' -> (R,G,B) in 0~1
    if argb_hex_string.startswith("#"):
        argb_hex_string = argb_hex_string[1:]
    if len(argb_hex_string) != 8:
        raise ValueError("Invalid argb hex string format. Expected #AARRGGBB.")
    r = int(argb_hex_string[2:4], 16) / 255.0
    g = int(argb_hex_string[4:6], 16) / 255.0
    b = int(argb_hex_string[6:8], 16) / 255.0
    return (r, g, b)

def get_excel_mapping(xlsx_path, sheet_name):
    # Returns { system_type_lower: (r,g,b) }
    workbook = load_workbook(xlsx_path)
    worksheet = workbook[sheet_name]
    start_row = None
    name_column = None
    color_column = None

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value == "Color":
                start_row = cell.row + 1
                name_column = cell.column - 1
                color_column = cell.column
                break
        if start_row is not None:
            break

    mapping = {}
    if start_row is None or name_column is None or color_column is None:
        print("'Color' header not found.")
        return mapping

    for i in range(start_row, worksheet.max_row + 1):
        name_cell = worksheet.cell(row=i, column=name_column)
        color_cell = worksheet.cell(row=i, column=color_column)
        name = (name_cell.value or "").strip()
        argb_hex = color_cell.fill.start_color.rgb
        if name and argb_hex:
            r, g, b = argb_to_rgb(argb_hex)
            mapping[str(name).lower()] = (r, g, b)
    return mapping

def get_body_items(prod):
    # Return a list of representation items for the product's Body representation (or first rep).
    rep_count = getattr(prod, "Representation", None)
    if not rep_count or not getattr(rep_count, "Representations", None):
        return []

    # Prefer the 'Body' representation; otherwise take the first one
    reps = rep_count.Representations or []
    body = next((r for r in reps if getattr(r, "RepresentationIdentifier", None) == "Body"),
                (reps[0] if reps else None))
    if not body:
        return []

    items = getattr(body, "Items", None) or []
    return list(items)

def get_or_make_psa(ifc, cache, r, g, b, transparency=0.0):
    # Cache by RGB(+transparency) to reuse styles across objects
    key = (round(r, 6), round(g, 6), round(b, 6), round(transparency, 6))
    if key in cache:
        return cache[key]

    rgb   = ifc.create_entity("IfcColourRgb", None, r, g, b)
    render = ifc.create_entity(
        "IfcSurfaceStyleRendering",
        rgb,                                    # SurfaceColour
        transparency,                           # Transparency (0 = opaque)
        None, None, None, None, None, None,     # (optional reflectance params)
        "NOTDEFINED"                            # ReflectanceMethod
    )
    surf  = ifc.create_entity("IfcSurfaceStyle", None, "BOTH", [render])
    psa   = ifc.create_entity("IfcPresentationStyleAssignment", [surf])

    cache[key] = psa
    return psa

def ensure_psa(ifc, style_or_psa):
    # If it’s already an IfcPresentationStyleAssignment, return it.
    if hasattr(style_or_psa, "is_a") and style_or_psa.is_a("IfcPresentationStyleAssignment"):
        return style_or_psa
    # Otherwise create a new one
    return ifc.create_entity("IfcPresentationStyleAssignment", [style_or_psa])


def assign_style_to_item_instance(ifc, item, style):
    psa = ensure_psa(ifc, style)
    # Assign style to a specific representation item (instance-level) by creating/updating an IfcStyledItem on that item.
    # Some schemas expose inverse as StyledByItem (list-like); handle robustly
    existing = getattr(item, "StyledByItem", None)
    # If there’s already one IfcStyledItem attached, it reuses it rather than creating a duplicate.
    # Each geometry item only has one styled item
    if existing and len(existing) > 0 and existing[0]:
        styled = existing[0]
        # It reads the existing list of IfcPresentationStyle references from styled.Styles
        # Adds the new one if it’s not already included.
        styles = list(styled.Styles or [])
        if psa not in styles:
            styles.append(psa)
            styled.Styles = styles
        return styled
    else:
        # Create new IfcStyledItem bound to the item
        return ifc.create_entity("IfcStyledItem", Item=item, Styles=[psa], Name=None)


def main():
    ifc = IfcStore.get_file()
    if not ifc:
        raise RuntimeError("No IFC open.")

    mapping = get_excel_mapping(XLSX_PATH, SHEET_NAME)
    print(mapping)
    if not mapping:
        print("[WARN] No mapping; abort.")
        return

    style_cache = {}
    assigned_items, no_body, no_color = 0, 0, set()

    for obj in bpy.data.objects:
        entity = tool.Ifc.get_entity(obj)
        if not entity:
            continue

        # Only work on files that have 7-digits in Name
        if not re.search(r'\b\d{7}\b', (entity.Name or "")):
            continue
         
        # Get the values (color name) on 'System Type' for each object
        psets = ifcopenshell.util.element.get_psets(entity, psets_only=True) or {}
        system_type = (psets.get("Mechanical", {}) or {}).get("System Type")
        #print("\nname:", entity.Name)
        #print("system type value:", system_type)
        if not system_type:
            continue

        key = (str(system_type) or "").strip().lower()
        rgb = mapping.get(key)
        #print("lookup key:", key, "rgb:", rgb)
        if not rgb:
            no_color.add(system_type)
            continue

        r, g, b = rgb
        style = get_or_make_psa(ifc, style_cache, r, g, b, transparency=0.0)
        #print("style:", style)

        items = get_body_items(entity)
        if not items:
            no_body += 1
            continue

        for item in items:
            assign_style_to_item_instance(ifc, item, style)
            assigned_items += 1
            
    print(f"[DONE] Styled {assigned_items} representation items (true instance-level).")
    if no_body:
        print(f"[INFO] {no_body} products had no Body representation or items.")
    if no_color:
        print(f"[INFO] Missing Excel colors for System Types: {sorted(set(no_color))}")
    print("[NOTE] Save the IFC to persist styles.")

if __name__ == "__main__":
    main()